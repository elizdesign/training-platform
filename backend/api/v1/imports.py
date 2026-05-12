# backend/api/v1/imports.py
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from sqlalchemy.orm import Session
from openpyxl import load_workbook
from datetime import date

from backend.db.session import get_db
from backend.models.user import User
from backend.models.client import Client
from backend.models.progress import Measurement
from backend.services.auth import get_current_user

router = APIRouter(prefix="/imports", tags=["Imports"])


@router.post("/excel")
def import_excel(
        file: UploadFile = File(...),
        dry_run: bool = Query(False, description="Тестовый прогон без сохранения в БД"),
        db: Session = Depends(get_db),
        current_user: User = Depends(get_current_user)
):
    # 1. Валидация файла
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Только .xlsx файлы")

    content = file.file.read()
    wb = load_workbook(io.BytesIO(content))

    report = {"clients": {"created": 0, "errors": []},
              "measurements": {"created": 0, "errors": []},
              "workouts": {"created": 0, "errors": []}}

    # 2. Парсинг клиентов
    if "Клиенты" in wb.sheetnames:
        ws = wb["Клиенты"]
        clients_map = {}  # phone -> client_id

        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                first_name, last_name, phone, age, goal = row
                if not phone:
                    report["clients"]["errors"].append("Пустой телефон")
                    continue

                # Поиск существующего или создание
                client = db.query(Client).filter(Client.phone == phone).first()
                if not client:
                    if dry_run:
                        clients_map[phone] = "new"
                        report["clients"]["created"] += 1
                        continue

                    client = Client(
                        first_name=first_name, last_name=last_name,
                        phone=phone, age=age or 0, goal=goal or "",
                        trainer_id=current_user.id
                    )
                    db.add(client)
                    db.flush()
                    clients_map[phone] = client.id
                    report["clients"]["created"] += 1
                else:
                    clients_map[phone] = client.id
            except Exception as e:
                report["clients"]["errors"].append(f"Строка {ws.max_row}: {str(e)}")

    # 3. Парсинг измерений
    if "Измерения" in wb.sheetnames:
        ws = wb["Измерения"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                phone, recorded_at, weight, waist, hips, chest = row
                client_id = clients_map.get(phone)
                if not client_id and not dry_run:
                    report["measurements"]["errors"].append(f"Клиент {phone} не найден")
                    continue

                m = Measurement(
                    client_id=client_id or 0,  # 0 для dry_run
                    recorded_at=recorded_at or date.today(),
                    weight=weight, waist=waist, hips=hips, chest=chest
                )
                if not dry_run:
                    db.add(m)
                report["measurements"]["created"] += 1
            except Exception as e:
                report["measurements"]["errors"].append(f"Ошибка: {str(e)}")

    # 4. Парсинг тренировок (упрощённо)
    if "Тренировки" in wb.sheetnames:
        ws = wb["Тренировки"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            try:
                title, workout_date, phone = row
                client_id = clients_map.get(phone)
                if not client_id and not dry_run:
                    report["workouts"]["errors"].append(f"Клиент {phone} не найден")
                    continue

                # Заглушка: создаём тренировку без упражнений для демо
                from backend.models.workout import Workout
                if not dry_run:
                    w = Workout(title=title, workout_date=workout_date or date.today(),
                                client_id=client_id or 0, trainer_id=current_user.id)
                    db.add(w)
                report["workouts"]["created"] += 1
            except Exception as e:
                report["workouts"]["errors"].append(f"Ошибка: {str(e)}")

    # 5. Транзакционность
    if not dry_run:
        try:
            db.commit()
        except Exception:
            db.rollback()
            raise HTTPException(status_code=500, detail="Критическая ошибка БД. Импорт отменён.")

    return {"dry_run": dry_run, "report": report}